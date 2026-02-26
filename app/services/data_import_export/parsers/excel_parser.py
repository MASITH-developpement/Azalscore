"""
AZALSCORE - Excel Parser
Parser Excel (xlsx)
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import Generator

from .base import BaseParser


class ExcelParser(BaseParser):
    """Parser Excel (xlsx)"""

    def parse(self, content: bytes, options: dict) -> Generator[dict, None, None]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl requis pour parser Excel: pip install openpyxl")

        sheet_name = options.get("sheet_name")
        sheet_index = options.get("sheet_index", 0)
        skip_rows = options.get("skip_rows", 0)
        has_header = options.get("has_header", True)

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[sheet_index]

        rows = ws.iter_rows(values_only=True)

        for _ in range(skip_rows):
            try:
                next(rows)
            except StopIteration:
                wb.close()
                return

        headers = []
        if has_header:
            try:
                header_row = next(rows)
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header_row)]
            except StopIteration:
                wb.close()
                return

        row_num = skip_rows + 2 if has_header else skip_rows + 1
        for row in rows:
            if not any(row):
                row_num += 1
                continue

            if headers:
                record = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    if isinstance(value, datetime):
                        record[header] = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, date):
                        record[header] = value.strftime("%Y-%m-%d")
                    else:
                        record[header] = value
            else:
                record = {f"col_{i}": val for i, val in enumerate(row)}

            record["__row_number__"] = row_num
            yield record
            row_num += 1

        wb.close()

    def get_headers(self, content: bytes, options: dict) -> list[str]:
        try:
            import openpyxl
        except ImportError:
            return []

        sheet_index = options.get("sheet_index", 0)
        skip_rows = options.get("skip_rows", 0)

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.worksheets[sheet_index]

        rows = ws.iter_rows(values_only=True)
        for _ in range(skip_rows):
            try:
                next(rows)
            except StopIteration:
                wb.close()
                return []

        try:
            header_row = next(rows)
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header_row)]
        except StopIteration:
            headers = []

        wb.close()
        return headers
