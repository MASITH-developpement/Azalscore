"""
AZALSCORE - Excel Exporter
Exporter Excel (xlsx)
"""
from __future__ import annotations

import io
from datetime import datetime, date
from decimal import Decimal

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Exporter Excel (xlsx)"""

    def export(self, data: list[dict], options: dict) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl requis pour export Excel")

        if not data:
            return b""

        sheet_name = options.get("sheet_name", "Export")
        columns = options.get("columns")
        style_header = options.get("style_header", True)
        auto_width = options.get("auto_width", True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if columns:
            headers = columns
        else:
            headers = [k for k in data[0].keys() if not k.startswith("__")]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            if style_header:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        for row_num, record in enumerate(data, start=2):
            for col_num, header in enumerate(headers, start=1):
                value = record.get(header)

                if isinstance(value, datetime):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.number_format = "DD/MM/YYYY HH:MM:SS"
                elif isinstance(value, date):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.number_format = "DD/MM/YYYY"
                elif isinstance(value, Decimal):
                    cell = ws.cell(row=row_num, column=col_num, value=float(value))
                    cell.number_format = "#,##0.00"
                else:
                    cell = ws.cell(row=row_num, column=col_num, value=value)

                cell.border = thin_border

        if auto_width:
            for col_num, header in enumerate(headers, start=1):
                max_length = len(str(header))
                for row in range(2, len(data) + 2):
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
